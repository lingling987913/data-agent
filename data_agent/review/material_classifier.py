"""
Review-Plus 材料角色分类服务

确定性优先策略：文件扩展名 + 文件名关键词 → xlsx 表头检测 → unknown
"""

import logging
from pathlib import Path

from data_agent.review.schemas import ReviewPlusMaterialRole, ReviewPlusCheckItem

logger = logging.getLogger(__name__)

_REVIEW_RULE_NAME_KEYWORDS = [
    "检查项", "审查标准", "评审准则", "评审标准",
    "审查准则", "评估标准", "文档检查需求", "检查需求", "准则",
    "review_rule", "rule", "criteria", "audit",
]

_CHECKLIST_NAME_KEYWORDS = [
    "检查单", "检查清单", "检查表", "产品保证工作检查单", "checklist",
]

_TASK_BOOK_NAME_KEYWORDS = [
    "任务书", "研制任务书", "任务要求", "合同要求", "task book", "statement of work",
]

_SUBJECT_REPORT_NAME_KEYWORDS = [
    "可靠性安全性设计与分析报告", "设计与分析报告", "分析报告", "设计报告",
    "报告", "方案", "设计", "report", "analysis report", "design report",
]

_SUBJECT_DOCUMENT_NAME_KEYWORDS = [
    "任务书", "需求", "设计", "报告", "方案", "分析",
    "说明书", "规范", "技术条件",
    "spec", "specification", "design", "requirements", "report", "solution",
]

_REVIEW_RULE_CONTENT_KEYWORDS = [
    "CHECK-", "审查准则", "检查项", "Severity",
    "severity: critical", "severity: major",
]

_CHECKLIST_CONTENT_KEYWORDS = [
    "产品保证工作检查单", "检查单", "检查结论", "检查内容", "检查要求",
]

_TASK_BOOK_CONTENT_KEYWORDS = [
    "研制任务书", "任务书", "任务要求", "交付物", "验收准则",
]

_SUBJECT_REPORT_CONTENT_KEYWORDS = [
    "可靠性安全性设计与分析", "可靠性分析", "安全性分析", "FMEA", "FTA", "SCA",
    "关键项目", "可靠性预计", "环境适应性",
]

_SUBJECT_DOCUMENT_CONTENT_KEYWORDS = [
    "REQ-GNC", "REQ-TOP", "设计方案", "需求规格", "DES-GNC", "SIM-GNC",
]

_XLSX_REVIEW_RULE_HEADERS = {
    "检查项目", "检查要求", "检查对象", "是否符合", "检查内容",
    "审查项目", "审查要求", "审查内容",
    "序号", "检查项",
}


def classify_material(
    name: str,
    content: str,
    file_path: str = "",
) -> tuple[ReviewPlusMaterialRole, float, str]:
    """对单个材料进行角色分类。

    Returns:
        (role, confidence, reason)
    """
    ext = Path(name).suffix.lower()
    name_lower = name.lower()

    # ── 1. xlsx/xls + 审查关键词文件名 → review_rule ──
    if ext in (".xlsx", ".xls"):
        if any(kw in name_lower for kw in _REVIEW_RULE_NAME_KEYWORDS):
            return (
                ReviewPlusMaterialRole.REVIEW_RULE,
                0.95,
                f"xlsx 文件名包含审查关键词",
            )
        # xlsx 内容检测：解析表头
        if file_path and Path(file_path).exists():
            header_role = _detect_xlsx_header(file_path)
            if header_role:
                return header_role

    # ── 2. 文件名匹配细分业务角色 ──
    for kw in _CHECKLIST_NAME_KEYWORDS:
        if kw.lower() in name_lower:
            return (
                ReviewPlusMaterialRole.CHECKLIST,
                0.9,
                f"文件名包含检查单关键词「{kw}」",
            )

    for kw in _TASK_BOOK_NAME_KEYWORDS:
        if kw.lower() in name_lower:
            return (
                ReviewPlusMaterialRole.TASK_BOOK,
                0.9,
                f"文件名包含任务书关键词「{kw}」",
            )

    for kw in _SUBJECT_REPORT_NAME_KEYWORDS:
        if kw.lower() in name_lower:
            return (
                ReviewPlusMaterialRole.SUBJECT_REPORT,
                0.88,
                f"文件名包含报告关键词「{kw}」",
            )

    # ── 3. 任意文件名匹配审查规则关键词 ──
    for kw in _REVIEW_RULE_NAME_KEYWORDS:
        if kw.lower() in name_lower:
            return (
                ReviewPlusMaterialRole.REVIEW_RULE,
                0.85,
                f"文件名包含关键词「{kw}」",
            )

    # ── 4. 任意文件名匹配待审文档关键词 ──
    for kw in _SUBJECT_DOCUMENT_NAME_KEYWORDS:
        if kw.lower() in name_lower:
            return (
                ReviewPlusMaterialRole.SUBJECT_DOCUMENT,
                0.8,
                f"文件名包含关键词「{kw}」",
            )

    # ── 5. 内容关键词匹配 ──
    content_sample = (content or "")[:2000]
    if content_sample:
        for kw in _CHECKLIST_CONTENT_KEYWORDS:
            if kw in content_sample:
                return (
                    ReviewPlusMaterialRole.CHECKLIST,
                    0.82,
                    f"内容包含检查单关键词「{kw}」",
                )
        for kw in _TASK_BOOK_CONTENT_KEYWORDS:
            if kw in content_sample:
                return (
                    ReviewPlusMaterialRole.TASK_BOOK,
                    0.82,
                    f"内容包含任务书关键词「{kw}」",
                )
        for kw in _SUBJECT_REPORT_CONTENT_KEYWORDS:
            if kw in content_sample:
                return (
                    ReviewPlusMaterialRole.SUBJECT_REPORT,
                    0.82,
                    f"内容包含报告关键词「{kw}」",
                )
        for kw in _REVIEW_RULE_CONTENT_KEYWORDS:
            if kw in content_sample:
                return (
                    ReviewPlusMaterialRole.REVIEW_RULE,
                    0.75,
                    f"内容包含审查规则关键词「{kw}」",
                )
        for kw in _SUBJECT_DOCUMENT_CONTENT_KEYWORDS:
            if kw in content_sample:
                return (
                    ReviewPlusMaterialRole.SUBJECT_DOCUMENT,
                    0.70,
                    f"内容包含待审文档关键词「{kw}」",
                )

    # ── 6. unknown ──
    return (
        ReviewPlusMaterialRole.UNKNOWN,
        0.3,
        "未匹配到已知角色模式",
    )


def _detect_xlsx_header(
    file_path: str,
) -> tuple[ReviewPlusMaterialRole, float, str] | None:
    """检查 xlsx 表头是否包含检查单特征列。"""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue

            headers = {str(cell).strip() for cell in first_row if cell}
            match_count = len(headers & _XLSX_REVIEW_RULE_HEADERS)
            if match_count >= 2:
                wb.close()
                return (
                    ReviewPlusMaterialRole.REVIEW_RULE,
                    0.9,
                    f"xlsx 表头匹配检查单模式 (sheet={sheet_name}, 匹配 {match_count} 列)",
                )
        wb.close()
    except Exception as e:
        logger.warning(f"[MaterialClassifier] xlsx 表头检测失败: {e}")
    return None


def classify_all_materials(
    materials: list[dict],
) -> list[tuple[ReviewPlusMaterialRole, float, str]]:
    """批量分类材料列表。"""
    results = []
    for mat in materials:
        role, confidence, reason = classify_material(
            name=mat.get("name", ""),
            content=mat.get("content", ""),
            file_path=mat.get("file_path", ""),
        )
        results.append((role, confidence, reason))
    return results
