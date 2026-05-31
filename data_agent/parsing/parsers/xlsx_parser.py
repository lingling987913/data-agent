"""Excel (.xlsx/.xls) format parser."""

from __future__ import annotations

import uuid

from data_agent.parsing.ingestion import markdown_table as _markdown_table
from data_agent.parsing.schemas import ParsedDocument, ParsedDocumentBlock


def parse_xlsx(file_path: str, parsed_doc: ParsedDocument) -> None:
    parsed_doc.parser_name = "openpyxl"
    parsed_doc.file_type = "attachment"
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        blocks: list[ParsedDocumentBlock] = []
        for sheet_name in workbook.sheetnames[:10]:
            sheet = workbook[sheet_name]
            rows: list[list[str]] = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx >= 500:
                    break
                rows.append([str(cell).strip() if cell is not None else "" for cell in row])
            if not rows:
                continue
            blocks.append(
                ParsedDocumentBlock(
                    block_id=str(uuid.uuid4()),
                    block_type="heading",
                    text=f"Sheet: {sheet_name}",
                    level=2,
                    order_index=len(blocks),
                )
            )
            width = max(len(row) for row in rows)
            padded = [row + [""] * (width - len(row)) for row in rows]
            blocks.append(
                ParsedDocumentBlock(
                    block_id=str(uuid.uuid4()),
                    block_type="table",
                    text=_markdown_table(padded[0], padded[1:] if len(padded) > 1 else []),
                    order_index=len(blocks),
                )
            )
        workbook.close()
        parsed_doc.blocks = blocks
        if not blocks:
            parsed_doc.parse_status = "degraded"
            parsed_doc.warnings.append("Excel workbook contained no readable rows.")
    except Exception as exc:
        parsed_doc.parse_status = "failed"
        parsed_doc.warnings.append(f"Excel extraction failed: {exc}")
