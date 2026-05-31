"""Lite material preview extraction for L0/L1 classification before full parsing."""

from __future__ import annotations

import csv
import subprocess
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET


def _truncate(text: str, max_chars: int) -> str:
    cleaned = (text or "").strip()
    if len(cleaned) <= max_chars:
        return cleaned
    return cleaned[:max_chars]


def _read_text_preview(file_path: str, max_chars: int) -> str:
    path = Path(file_path)
    if not path.exists():
        return ""
    raw = path.read_bytes()[: max_chars * 4]
    for enc in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return _truncate(raw.decode(enc), max_chars)
        except Exception:
            continue
    return ""


def _docx_preview(file_path: str, max_chars: int) -> str:
    try:
        with zipfile.ZipFile(file_path) as zf:
            xml_bytes = zf.read("word/document.xml")
        root = ET.fromstring(xml_bytes)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        parts: list[str] = []
        for paragraph in root.findall(".//w:p", ns):
            texts = [node.text for node in paragraph.findall(".//w:t", ns) if node.text]
            line = "".join(texts).strip()
            if line:
                parts.append(line)
            if sum(len(part) + 1 for part in parts) >= max_chars:
                break
        return _truncate("\n".join(parts), max_chars)
    except Exception:
        return ""


def _pdf_preview(file_path: str, max_chars: int) -> str:
    try:
        result = subprocess.run(
            ["pdftotext", "-f", "1", "-l", "2", file_path, "-"],
            capture_output=True,
            text=True,
            timeout=15,
            check=False,
        )
        if result.returncode == 0 and result.stdout.strip():
            return _truncate(result.stdout, max_chars)
    except Exception:
        pass
    return ""


def _excel_preview(file_path: str, max_chars: int) -> str:
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in workbook.sheetnames[:2]:
            sheet = workbook[sheet_name]
            rows: list[str] = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx >= 20:
                    break
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                parts.append(f"[{sheet_name}] " + " ; ".join(rows))
        workbook.close()
        return _truncate("\n".join(parts), max_chars)
    except Exception:
        return ""


def _csv_preview(file_path: str, max_chars: int) -> str:
    text = _read_text_preview(file_path, max_chars)
    if not text:
        return ""
    try:
        rows = list(csv.reader(text.splitlines()))[:20]
        return _truncate("\n".join(" | ".join(row) for row in rows if row), max_chars)
    except Exception:
        return _truncate(text, max_chars)


def extract_material_preview(
    file_path: str,
    file_name: str,
    *,
    max_chars: int = 2000,
) -> str:
    """Extract a lightweight text preview for TaskClassifier without full parsing."""
    ext = Path(file_name).suffix.lower()
    if ext in (".txt", ".md"):
        return _read_text_preview(file_path, max_chars)
    if ext == ".docx":
        return _docx_preview(file_path, max_chars)
    if ext == ".pdf":
        return _pdf_preview(file_path, max_chars)
    if ext in (".xlsx", ".xls"):
        return _excel_preview(file_path, max_chars)
    if ext == ".csv":
        return _csv_preview(file_path, max_chars)
    return ""
