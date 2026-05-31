from __future__ import annotations

from data_agent.domain.material_roles import ReviewCheckItem


def extract_review_check_items(file_path: str) -> list[ReviewCheckItem]:
    """从 XLSX 检查需求 Sheet1 确定性抽取检查项。"""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX parsing") from exc

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    items: list[ReviewCheckItem] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        cells = [str(c).strip() if c is not None else "" for c in row]
        while len(cells) < 7:
            cells.append("")
        item_no, _, check_subject, check_target, requirement, _, remark = cells[:7]
        if not check_subject and not requirement:
            continue
        items.append(
            ReviewCheckItem(
                item_no=item_no or str(row_idx - 1),
                check_subject=check_subject,
                check_target=check_target,
                requirement=requirement,
                remark=remark,
            )
        )
    workbook.close()
    return items
